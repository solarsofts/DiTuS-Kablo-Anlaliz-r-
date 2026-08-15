from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import shlex
import signal
import subprocess
import sys
import tempfile
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Iterable, Sequence

from tools.publish_integrity_audit import (
    IGNORED_DIR_NAMES,
    scan_package,
)


ACCEPTANCE_SCHEMA_VERSION = "1.0"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _load_engine_baseline(root: Path, baseline_path: Path) -> dict[str, str]:
    baseline: dict[str, str] = {}
    for line_number, raw in enumerate(baseline_path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw.strip()
        if not line:
            continue
        try:
            digest, relative = line.split(None, 1)
        except ValueError as exc:
            raise ValueError(f"Invalid engine baseline line {line_number}") from exc
        relative = relative.strip()
        if len(digest) != 64 or any(char not in "0123456789abcdef" for char in digest.lower()):
            raise ValueError(f"Invalid SHA-256 at engine baseline line {line_number}")
        baseline[relative] = digest.lower()
    return baseline


def verify_engine_lock(root: Path, baseline_path: Path) -> dict[str, Any]:
    expected = _load_engine_baseline(root, baseline_path)
    current_paths: set[str] = set()
    mismatches: list[dict[str, str]] = []
    for folder in (root / "src/ucd/calculations", root / "src/ucd/models"):
        for path in sorted(item for item in folder.rglob("*") if item.is_file() and "__pycache__" not in item.parts):
            relative = path.relative_to(root).as_posix()
            current_paths.add(relative)
            actual = _sha256(path)
            expected_digest = expected.get(relative)
            if expected_digest != actual:
                mismatches.append(
                    {
                        "path": relative,
                        "expected_sha256": expected_digest or "<missing>",
                        "actual_sha256": actual,
                    }
                )
    missing = sorted(set(expected) - current_paths)
    unexpected = sorted(current_paths - set(expected))
    status = "PASS" if not mismatches and not missing and not unexpected else "FAIL"
    return {
        "status": status,
        "baseline_file": baseline_path.relative_to(root).as_posix(),
        "baseline_file_sha256": _sha256(baseline_path),
        "expected_file_count": len(expected),
        "verified_file_count": len(current_paths),
        "mismatches": mismatches,
        "missing_paths": missing,
        "unexpected_paths": unexpected,
    }


def _parse_junit(path: Path, exit_code: int, elapsed_seconds: float) -> dict[str, Any]:
    root = ET.parse(path).getroot()
    suites = [root] if root.tag == "testsuite" else list(root.findall("testsuite"))
    totals = {"tests": 0, "failures": 0, "errors": 0, "skipped": 0}
    failed_cases: list[str] = []
    for suite in suites:
        for key in totals:
            totals[key] += int(float(suite.attrib.get(key, "0")))
        for case in suite.findall(".//testcase"):
            if case.find("failure") is not None or case.find("error") is not None:
                class_name = case.attrib.get("classname", "")
                name = case.attrib.get("name", "")
                failed_cases.append(f"{class_name}::{name}".strip(":"))
    status = "PASS" if exit_code == 0 and totals["failures"] == 0 and totals["errors"] == 0 else "FAIL"
    return {
        "status": status,
        "exit_code": exit_code,
        "tests": totals["tests"],
        "passed": totals["tests"] - totals["failures"] - totals["errors"] - totals["skipped"],
        "failures": totals["failures"],
        "errors": totals["errors"],
        "skipped": totals["skipped"],
        "elapsed_seconds": round(elapsed_seconds, 3),
        "failed_cases": failed_cases,
    }


def _test_files(root: Path) -> list[Path]:
    return sorted((root / "tests").glob("test_*.py"))


def _files_for_shard(root: Path, shard_index: int, shard_count: int) -> list[Path]:
    if shard_count < 1:
        raise ValueError("shard_count must be at least 1")
    if shard_index < 0 or shard_index >= shard_count:
        raise ValueError("shard_index is outside shard_count")
    files = _test_files(root)
    return [path for index, path in enumerate(files) if index % shard_count == shard_index]


def _shard_directory(root: Path) -> Path:
    path = root / ".release_acceptance" / "pytest_shards"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _shard_summary_path(root: Path, shard_index: int, shard_count: int) -> Path:
    return _shard_directory(root) / f"shard_{shard_index + 1:02d}_of_{shard_count:02d}.json"


def _run_pytest_selection(
    root: Path,
    selections: list[str],
    *,
    timeout_seconds: float = 180.0,
) -> dict[str, Any]:
    """Run one isolated pytest selection and return parsed JUnit evidence.

    On POSIX, pytest is wrapped by a tiny shell that writes the parent exit code
    to a file. Numerical solver descendants may outlive pytest and keep process
    descriptors open; the acceptance parent therefore watches the result file,
    then cleans the complete process group without waiting for descendant EOF.
    Windows keeps the direct child path and process-level timeout.
    """
    with tempfile.TemporaryDirectory(prefix="ditus-pytest-") as temporary:
        temp_root = Path(temporary)
        junit_path = temp_root / "pytest-junit.xml"
        output_path = temp_root / "pytest-output.txt"
        exit_path = temp_root / "pytest-exit.txt"
        env = os.environ.copy()
        existing = env.get("PYTHONPATH", "")
        env["PYTHONPATH"] = os.pathsep.join(
            [str(root / "src"), str(root)] + ([existing] if existing else [])
        )
        command = [
            sys.executable,
            "-m",
            "pytest",
            "-q",
            *selections,
            f"--junitxml={junit_path}",
        ]
        started = time.monotonic()
        timed_out = False

        if os.name != "nt":
            script = (
                f"{shlex.join(command)} > {shlex.quote(str(output_path))} 2>&1; "
                "code=$?; "
                f"printf '%s\\n' \"$code\" > {shlex.quote(str(exit_path))}; "
                "exit $code"
            )
            process = subprocess.Popen(
                ["bash", "-c", script],
                cwd=root,
                env=env,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
            deadline = started + timeout_seconds
            while not exit_path.exists() and time.monotonic() < deadline:
                time.sleep(0.05)
            if exit_path.exists():
                try:
                    exit_code = int(exit_path.read_text(encoding="utf-8").strip())
                except (OSError, ValueError):
                    exit_code = int(process.poll() or 1)
            else:
                timed_out = True
                exit_code = 124
            try:
                os.killpg(process.pid, signal.SIGTERM)
                time.sleep(0.05)
                os.killpg(process.pid, signal.SIGKILL)
            except ProcessLookupError:
                pass
            try:
                process.wait(timeout=10)
            except subprocess.TimeoutExpired:
                pass
        else:
            with output_path.open("w", encoding="utf-8", errors="replace") as output_stream:
                process = subprocess.Popen(
                    command,
                    cwd=root,
                    env=env,
                    stdout=output_stream,
                    stderr=subprocess.STDOUT,
                    text=True,
                )
                try:
                    process.wait(timeout=timeout_seconds)
                except subprocess.TimeoutExpired:
                    timed_out = True
                    process.kill()
                    process.wait(timeout=10)
            exit_code = 124 if timed_out else int(process.returncode or 0)

        elapsed = time.monotonic() - started
        output = output_path.read_text(encoding="utf-8", errors="replace") if output_path.exists() else ""
        if timed_out:
            return {
                "status": "TIMEOUT",
                "exit_code": 124,
                "tests": 0,
                "passed": 0,
                "failures": 0,
                "errors": 1,
                "skipped": 0,
                "elapsed_seconds": round(elapsed, 3),
                "failed_cases": [f"timeout: {' '.join(selections)}"],
                "output_tail": output[-4000:],
            }
        if not junit_path.exists():
            return {
                "status": "FAIL",
                "exit_code": exit_code or 1,
                "tests": 0,
                "passed": 0,
                "failures": 0,
                "errors": 1,
                "skipped": 0,
                "elapsed_seconds": round(elapsed, 3),
                "failed_cases": ["pytest did not produce JUnit XML"],
                "output_tail": output[-4000:],
            }
        result = _parse_junit(junit_path, exit_code, elapsed)
        result["output_tail"] = output[-4000:] if result["status"] != "PASS" else ""
        return result


def _collect_test_nodes(root: Path, relative_file: str) -> list[str]:
    env = os.environ.copy()
    existing = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = os.pathsep.join(
        [str(root / "src"), str(root)] + ([existing] if existing else [])
    )
    completed = subprocess.run(
        [sys.executable, "-m", "pytest", "--collect-only", "-q", relative_file],
        cwd=root,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
        timeout=60,
    )
    if completed.returncode != 0:
        return []
    prefix = relative_file + "::"
    return [line.strip() for line in completed.stdout.splitlines() if line.strip().startswith(prefix)]


def _run_selection_in_fresh_parent(
    root: Path,
    selection: str,
    *,
    timeout_seconds: float = 180.0,
) -> dict[str, Any]:
    """Execute one selection through a disposable process group.

    Some numerical/UI tests can leave descendants alive after the disposable
    Python parent has written its JSON result. ``subprocess.run`` may then wait
    on inherited descriptors in containerized acceptance environments. Use an
    explicit session, file-backed output and unconditional process-group cleanup.
    """
    with tempfile.TemporaryDirectory(prefix="ditus-pytest-parent-") as temporary:
        temp_root = Path(temporary)
        result_path = temp_root / "result.json"
        output_path = temp_root / "parent-output.txt"
        env = os.environ.copy()
        existing = env.get("PYTHONPATH", "")
        env["PYTHONPATH"] = os.pathsep.join(
            [str(root / "src"), str(root)] + ([existing] if existing else [])
        )
        command = [
            sys.executable,
            "-m",
            "tools.run_isolated_pytest",
            "--root",
            str(root),
            "--selection",
            selection,
            "--output",
            str(result_path),
            "--timeout",
            str(timeout_seconds),
        ]
        timed_out = False
        with output_path.open("w", encoding="utf-8", errors="replace") as output_stream:
            process = subprocess.Popen(
                command,
                cwd=root,
                env=env,
                stdout=output_stream,
                stderr=subprocess.STDOUT,
                text=True,
                start_new_session=(os.name != "nt"),
            )
            try:
                process.wait(timeout=timeout_seconds + 30.0)
            except subprocess.TimeoutExpired:
                timed_out = True
            finally:
                if process.poll() is None:
                    if os.name == "nt":
                        process.kill()
                    else:
                        try:
                            os.killpg(process.pid, signal.SIGKILL)
                        except ProcessLookupError:
                            pass
                    process.wait(timeout=10)
                elif os.name != "nt":
                    try:
                        os.killpg(process.pid, signal.SIGTERM)
                        time.sleep(0.05)
                        os.killpg(process.pid, signal.SIGKILL)
                    except ProcessLookupError:
                        pass
        if result_path.exists():
            return json.loads(result_path.read_text(encoding="utf-8"))
        output = output_path.read_text(encoding="utf-8", errors="replace") if output_path.exists() else ""
        return {
            "status": "TIMEOUT" if timed_out else "FAIL",
            "exit_code": 124 if timed_out else int(process.returncode or 1),
            "tests": 0,
            "passed": 0,
            "failures": 0,
            "errors": 1,
            "skipped": 0,
            "elapsed_seconds": 0.0,
            "failed_cases": [
                ("isolated parent timed out: " if timed_out else "isolated parent did not produce result: ")
                + selection
            ],
            "selection": selection,
            "output_tail": output[-4000:],
        }


def _aggregate_isolated_results(results: list[dict[str, Any]], files: list[str]) -> dict[str, Any]:
    totals = {
        key: sum(int(item.get(key, 0)) for item in results)
        for key in ("tests", "passed", "failures", "errors", "skipped")
    }
    failed_cases = [case for item in results for case in item.get("failed_cases", [])]
    status = "PASS" if all(item.get("status") == "PASS" for item in results) else "FAIL"
    return {
        "status": status,
        "exit_code": 0 if status == "PASS" else 1,
        **totals,
        "elapsed_seconds": round(sum(float(item.get("elapsed_seconds", 0.0)) for item in results), 3),
        "failed_cases": failed_cases,
        "files": files,
        "process_isolated_files": True,
        "runs": results,
    }


def run_pytest_files(root: Path, files: list[Path]) -> dict[str, Any]:
    relative_files = [path.relative_to(root).as_posix() for path in files]
    if not files:
        return {
            "status": "PASS",
            "exit_code": 0,
            "tests": 0,
            "passed": 0,
            "failures": 0,
            "errors": 0,
            "skipped": 0,
            "elapsed_seconds": 0.0,
            "failed_cases": [],
            "files": [],
            "process_isolated_files": True,
            "runs": [],
        }

    isolated_results: list[dict[str, Any]] = []
    node_isolated_files = {
        "tests/test_catalog_selection.py",
        "tests/test_shadow_validation.py",
        "tests/test_thermal_optimization.py",
        "tests/test_workflow_run_registry.py",
    }
    execution_files = sorted(relative_files, key=lambda item: (item not in node_isolated_files, item))
    for relative_file in execution_files:
        if relative_file in node_isolated_files:
            nodes = _collect_test_nodes(root, relative_file)
            node_results: list[dict[str, Any]] = []
            for node in nodes:
                node_result = _run_selection_in_fresh_parent(root, node)
                node_result["selection"] = node
                node_results.append(node_result)
            fallback = _aggregate_isolated_results(node_results, [relative_file])
            fallback["selection"] = relative_file
            fallback["node_isolation_fallback"] = True
            isolated_results.append(fallback)
            continue

        result = _run_pytest_selection(root, [relative_file])
        result["selection"] = relative_file
        if result.get("status") != "TIMEOUT":
            isolated_results.append(result)
            continue

        # A small number of legacy solver/UI files can deadlock only when their
        # tests share one interpreter. Preserve every test by retrying the exact
        # collected node IDs in independent processes.
        nodes = _collect_test_nodes(root, relative_file)
        if not nodes:
            isolated_results.append(result)
            continue
        node_results: list[dict[str, Any]] = []
        for node in nodes:
            node_result = _run_selection_in_fresh_parent(root, node)
            node_result["selection"] = node
            node_results.append(node_result)
        fallback = _aggregate_isolated_results(node_results, [relative_file])
        fallback["selection"] = relative_file
        fallback["node_isolation_fallback"] = True
        isolated_results.append(fallback)

    return _aggregate_isolated_results(isolated_results, relative_files)


def run_pytest_shard(root: Path, shard_index: int, shard_count: int) -> dict[str, Any]:
    result = run_pytest_files(root, _files_for_shard(root, shard_index, shard_count))
    result["shard_index"] = shard_index
    result["shard_count"] = shard_count
    path = _shard_summary_path(root, shard_index, shard_count)
    path.write_text(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def aggregate_pytest_shards(root: Path, shard_count: int) -> dict[str, Any]:
    shards: list[dict[str, Any]] = []
    missing: list[str] = []
    for shard_index in range(shard_count):
        path = _shard_summary_path(root, shard_index, shard_count)
        if not path.exists():
            missing.append(path.name)
            continue
        shards.append(json.loads(path.read_text(encoding="utf-8")))
    if missing:
        return {
            "status": "FAIL",
            "exit_code": 1,
            "tests": 0,
            "passed": 0,
            "failures": 0,
            "errors": len(missing),
            "skipped": 0,
            "elapsed_seconds": 0.0,
            "failed_cases": [f"missing shard: {name}" for name in missing],
            "shard_count": shard_count,
            "process_isolated_shards": True,
            "shards": shards,
        }
    totals = {key: sum(int(shard.get(key, 0)) for shard in shards) for key in ("tests", "passed", "failures", "errors", "skipped")}
    failed_cases = [case for shard in shards for case in shard.get("failed_cases", [])]
    status = "PASS" if all(shard.get("status") == "PASS" for shard in shards) else "FAIL"
    return {
        "status": status,
        "exit_code": 0 if status == "PASS" else 1,
        **totals,
        "elapsed_seconds": round(sum(float(shard.get("elapsed_seconds", 0.0)) for shard in shards), 3),
        "failed_cases": failed_cases,
        "shard_count": shard_count,
        "process_isolated_shards": True,
        "shards": shards,
    }


def run_pytest(root: Path, shard_count: int = 8) -> dict[str, Any]:
    shard_dir = _shard_directory(root)
    for stale in shard_dir.glob("shard_*_of_*.json"):
        stale.unlink()
    for shard_index in range(shard_count):
        run_pytest_shard(root, shard_index, shard_count)
    return aggregate_pytest_shards(root, shard_count)


def _iter_manifest_files(root: Path, exclusions: set[str]) -> Iterable[Path]:
    for path in sorted(item for item in root.rglob("*") if item.is_file()):
        relative = path.relative_to(root)
        rel = relative.as_posix()
        if rel in exclusions:
            continue
        if any(part in IGNORED_DIR_NAMES for part in relative.parts):
            continue
        yield path


def build_and_verify_manifest(root: Path, manifest_path: Path, exclusions: set[str]) -> dict[str, Any]:
    lines = []
    for path in _iter_manifest_files(root, exclusions | {manifest_path.relative_to(root).as_posix()}):
        lines.append(f"{_sha256(path)}  {path.relative_to(root).as_posix()}")
    manifest_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    mismatches: list[str] = []
    verified = 0
    for raw in manifest_path.read_text(encoding="utf-8").splitlines():
        if not raw.strip():
            continue
        digest, relative = raw.split(None, 1)
        candidate = root / relative.strip()
        if not candidate.is_file() or _sha256(candidate) != digest:
            mismatches.append(relative.strip())
        else:
            verified += 1
    return {
        "status": "PASS" if not mismatches else "FAIL",
        "manifest_file": manifest_path.relative_to(root).as_posix(),
        "manifest_sha256": _sha256(manifest_path),
        "entry_count": len(lines),
        "verified_count": verified,
        "mismatches": mismatches,
        "excluded_generated_attestations": sorted(exclusions),
    }



def audit_procurement_drum_plan(root: Path) -> dict[str, Any]:
    """Independently validate the packaged synthetic procurement plan."""
    try:
        from ucd.models.project import ProjectData
        from ucd.calculations.procurement import build_procurement_package
        source = root / "examples/synthetic_20km_procurement.ucd.json"
        json_output = root / "examples/synthetic_20km_boq_bom_rfq.latest.json"
        csv_output = root / "examples/synthetic_20km_boq_bom_rfq.latest_drum_plan.csv"
        xlsx_output = root / "examples/synthetic_20km_boq_bom_rfq.latest.xlsx"
        if not all(path.is_file() for path in (source, json_output, csv_output, xlsx_output)):
            return {"status": "FAIL", "violations": ["Packaged procurement source/output files are incomplete."]}
        project = ProjectData.from_dict(json.loads(source.read_text(encoding="utf-8")))
        package = build_procurement_package(project)
        plan = package.drum_plan
        violations: list[str] = []
        tolerance = 1.0e-6
        if plan.drum_plan_status != "VALID": violations.append(f"drum_plan_status={plan.drum_plan_status}")
        if plan.invalid_drum_count: violations.append(f"invalid_drum_count={plan.invalid_drum_count}")
        if plan.unassigned_cut_count: violations.append(f"unassigned_cut_count={plan.unassigned_cut_count}")
        if plan.overload_total_m > tolerance: violations.append(f"overload_total_m={plan.overload_total_m}")
        if plan.unallocated_total_m > tolerance: violations.append(f"unallocated_total_m={plan.unallocated_total_m}")
        if abs(plan.accounting_residual_m) > tolerance: violations.append(f"accounting_residual_m={plan.accounting_residual_m}")
        seen: list[str] = []
        for drum in package.drums:
            if drum.loaded_length_m > drum.maximum_length_m + tolerance:
                violations.append(f"{drum.drum_id}:loaded>{drum.maximum_length_m}")
            if drum.overload_m > tolerance:
                violations.append(f"{drum.drum_id}:overload={drum.overload_m}")
            seen.extend(cut.cut_id for cut in drum.cuts)
        if len(seen) != len(set(seen)): violations.append("duplicate_cut_assignments")
        if len(seen) != 126: violations.append(f"route_cut_count={len(seen)}")

        payload = json.loads(json_output.read_text(encoding="utf-8"))
        packaged_summary = payload.get("summary", {})
        for key, expected in {
            "drum_plan_status": plan.drum_plan_status,
            "drum_count": len(package.drums),
            "overload_total_m": plan.overload_total_m,
            "unallocated_total_m": plan.unallocated_total_m,
            "cross_bonding_link_box_units": 28,
            "grounding_link_box_units": 12,
            "svl_set_units": 28,
            "svl_pole_units": 84,
        }.items():
            if packaged_summary.get(key) != expected:
                violations.append(f"json_summary_mismatch:{key}")

        import csv as _csv
        with csv_output.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(_csv.DictReader(handle, delimiter=";"))
        if len(rows) != len(package.drums): violations.append("csv_drum_count_mismatch")
        csv_overload = sum(float(row.get("Aşım m", 0) or 0) for row in rows)
        if abs(csv_overload - plan.overload_total_m) > tolerance: violations.append("csv_overload_mismatch")

        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_output, read_only=True, data_only=False)
            ws = wb["Makara Planı"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            required = {"Aşım [m]", "Kapasite Bakiyesi [m]", "Sipariş/Fire Payı [m]"}
            if not required.issubset(set(headers)): violations.append("xlsx_drum_columns_missing")
            wb.close()
        except Exception as exc:
            violations.append(f"xlsx_validation_error:{type(exc).__name__}")

        accessory = package.accessory_plan
        if accessory.status != "VALID": violations.append(f"accessory_plan_status={accessory.status}")
        return {
            "status": "PASS" if not violations else "FAIL",
            "source_project": source.relative_to(root).as_posix(),
            "procurement_output_sha256": _sha256(json_output),
            "drum_plan_status": plan.drum_plan_status,
            "maximum_drum_length_m": max((d.maximum_length_m for d in package.drums), default=0.0),
            "drum_count": len(package.drums),
            "valid_drum_count": plan.valid_drum_count,
            "invalid_drum_count": plan.invalid_drum_count,
            "route_cut_count": len(seen),
            "unassigned_cut_count": plan.unassigned_cut_count,
            "route_cut_total_m": plan.route_cut_total_m,
            "order_allowance_total_m": plan.order_allowance_total_m,
            "spare_stock_total_m": plan.spare_stock_total_m,
            "order_quantity_total_m": package.summary.order_single_core_length_m,
            "allocated_total_m": plan.allocated_total_m,
            "unallocated_total_m": plan.unallocated_total_m,
            "overload_total_m": plan.overload_total_m,
            "maximum_single_drum_overload_m": plan.maximum_single_drum_overload_m,
            "accounting_residual_m": plan.accounting_residual_m,
            "duplicate_cut_assignments": len(seen) - len(set(seen)),
            "accessory_plan_status": accessory.status,
            "cross_bonding_link_box_units": package.summary.cross_bonding_link_box_units,
            "grounding_link_box_units": package.summary.grounding_link_box_units,
            "svl_set_units": package.summary.svl_set_units,
            "svl_pole_units": package.summary.svl_pole_units,
            "export_consistency": "PASS" if not any("mismatch" in item or "xlsx_" in item for item in violations) else "FAIL",
            "violations": violations,
        }
    except Exception as exc:
        return {"status": "FAIL", "violations": [f"{type(exc).__name__}: {exc}"]}

def _compose_payload(
    version: str,
    pytest_result: dict[str, Any],
    audit_result: Any,
    engine_lock: dict[str, Any],
    manifest: dict[str, Any],
    procurement_drum_plan: dict[str, Any],
    generated_output_self_check: str,
) -> dict[str, Any]:
    component_statuses = {
        "pytest": pytest_result["status"],
        "publish_integrity": audit_result.status,
        "engine_lock": engine_lock["status"],
        "manifest": manifest["status"],
        "procurement_drum_plan": procurement_drum_plan["status"],
        "generated_output_self_check": generated_output_self_check,
    }
    overall = "PASS" if all(value == "PASS" for value in component_statuses.values()) else "FAIL"
    return {
        "schema_version": ACCEPTANCE_SCHEMA_VERSION,
        "package_version": version,
        "overall_status": overall,
        "component_statuses": component_statuses,
        "pytest": pytest_result,
        "publish_integrity": audit_result.to_dict(),
        "engine_lock": engine_lock,
        "manifest": manifest,
        "procurement_drum_plan": procurement_drum_plan,
        "generated_output_self_check": generated_output_self_check,
        "historical_correction": {
            "superseded_documents": [
                "PACKAGED_TEST_RESULTS_v0.16.9.4.14.txt",
                "PUBLISH_CLEANUP_AUDIT_v0.16.9.4.14.md",
            ],
            "statement": "Earlier PDF and maximum-drum-length PASS claims were not supported by complete machine-readable evidence and are not accepted as evidence.",
        },
        "provenance": "Generated automatically by tools/run_release_acceptance.py; manual PASS editing is not authoritative.",
    }


def _render_text(payload: dict[str, Any]) -> str:
    audit = payload["publish_integrity"]
    stats = audit["stats"]
    tests = payload["pytest"]
    manifest = payload["manifest"]
    engine = payload["engine_lock"]
    procurement = payload["procurement_drum_plan"]
    return "\n".join(
        [
            f"DiTuS Kablo Analizör v{payload['package_version']}",
            "Otomatik paket kabul sonucu",
            "",
            f"GENEL SONUÇ: {payload['overall_status']}",
            f"Testler: {tests['status']} / {tests['passed']}/{tests['tests']} PASS / skipped={tests['skipped']}",
            f"Yayın veri bütünlüğü: {audit['status']} / blocker={audit['blocking_finding_count']} / allowed={audit['allowed_finding_count']}",
            f"PDF: {stats['pdf_files_scanned']}/{stats['pdf_files_seen']} dosya; {stats['pdf_pages_scanned']}/{stats['pdf_pages_seen']} sayfa; {stats['pdf_text_characters_scanned']} karakter",
            f"PDF metadata: {stats['pdf_metadata_fields_scanned']} alan; taranamayan nesne={stats['unscannable_objects']}",
            f"Hesap/model motor kilidi: {engine['status']} / {engine['verified_file_count']}/{engine['expected_file_count']}",
            f"Manifest: {manifest['status']} / {manifest['verified_count']}/{manifest['entry_count']}",
            f"Makara planı sayısal denetimi: {procurement['status']} / plan={procurement.get('drum_plan_status')} / makara={procurement.get('drum_count')} / aşım={procurement.get('overload_total_m', 0):.3f} m / atanmamış={procurement.get('unassigned_cut_count', 0)}",
            f"Bonding aksesuar denetimi: {procurement.get('accessory_plan_status')} / cross LB={procurement.get('cross_bonding_link_box_units')} / grounding LB={procurement.get('grounding_link_box_units')} / SVL set/pol={procurement.get('svl_set_units')}/{procurement.get('svl_pole_units')}",
            f"Üretilen kabul belgeleri öz-denetimi: {payload['generated_output_self_check']}",
            "",
            "Bu belge test ve denetim çıktısından otomatik üretilmiştir; elle yazılmış PASS beyanı değildir.",
            "v0.16.9.4.14 içindeki PDF tarama beyanı doğrulanmamış tarihsel kayıttır ve kanıt olarak kullanılmamıştır.",
        ]
    ) + "\n"


def _render_markdown(payload: dict[str, Any]) -> str:
    audit = payload["publish_integrity"]
    stats = audit["stats"]
    tests = payload["pytest"]
    manifest = payload["manifest"]
    engine = payload["engine_lock"]
    procurement = payload["procurement_drum_plan"]
    lines = [
        f"# Paket Kabul ve Yayın Veri Bütünlüğü — v{payload['package_version']}",
        "",
        f"## Genel sonuç: **{payload['overall_status']}**",
        "",
        "| Kapı | Sonuç | Kanıt |",
        "|---|---:|---|",
        f"| Pytest | {tests['status']} | {tests['passed']}/{tests['tests']} PASS, {tests['skipped']} skipped |",
        f"| Yayın veri bütünlüğü | {audit['status']} | {audit['blocking_finding_count']} blocker, {audit['allowed_finding_count']} konuma bağlı izin |",
        f"| Hesap/model motor kilidi | {engine['status']} | {engine['verified_file_count']}/{engine['expected_file_count']} dosya byte doğrulandı |",
        f"| Manifest | {manifest['status']} | {manifest['verified_count']}/{manifest['entry_count']} |",
        f"| Makara planı sayısal denetimi | {procurement['status']} | {procurement.get('drum_count')} makara, aşım {procurement.get('overload_total_m', 0):.3f} m, atanmamış {procurement.get('unassigned_cut_count', 0)} |",
        f"| Bonding aksesuar planı | {procurement.get('accessory_plan_status')} | Cross LB {procurement.get('cross_bonding_link_box_units')}, grounding LB {procurement.get('grounding_link_box_units')}, SVL set/pol {procurement.get('svl_set_units')}/{procurement.get('svl_pole_units')} |",
        f"| Kabul belgeleri öz-denetimi | {payload['generated_output_self_check']} | JSON/TXT/MD yeniden tarandı |",
        "",
        "## PDF kapsam kanıtı",
        "",
        f"- PDF dosyası: {stats['pdf_files_scanned']}/{stats['pdf_files_seen']}",
        f"- PDF sayfası: {stats['pdf_pages_scanned']}/{stats['pdf_pages_seen']}",
        f"- Çıkarılan PDF metni: {stats['pdf_text_characters_scanned']} karakter",
        f"- PDF metadata alanı: {stats['pdf_metadata_fields_scanned']}",
        f"- PDF annotation/form/ek: {stats['pdf_annotations_scanned']}/{stats['pdf_form_fields_scanned']}/{stats['pdf_attachments_scanned']}",
        f"- Taranamayan nesne: {stats['unscannable_objects']}",
        "",
        "## Denetim sözleşmesi",
        "",
        "Genel desenler ve kimlik metadata kontrolleri ana kapıdır. Geçmiş sızıntılar SHA-256 regresyon parmak izleriyle tamamlayıcı olarak izlenir. İzinler yalnız belirli kural ve belirli dosya yolunda geçerlidir. PDF açılamaz, şifreliyse veya metin çıkarımı boş kalırsa denetim fail-closed davranır.",
        "",
        "Yeni kabul belgeleri tek yapılandırılmış JSON sonucundan otomatik türetilmiştir. Manifest, kendisi ile sonradan üretilen kabul attestasyonlarını çevrimsel hash bağı nedeniyle kapsam dışında bırakır; bu üç attestasyon ayrıca yayın tarayıcısıyla öz-denetlenir.",
        "",
        "## Makara ve aksesuar kapsam kanıtı",
        "",
        f"- Sipariş miktarı: {procurement.get('order_quantity_total_m', 0):.3f} m",
        f"- Tahsis edilen / edilmeyen: {procurement.get('allocated_total_m', 0):.3f} / {procurement.get('unallocated_total_m', 0):.3f} m",
        f"- Güzergâh kesimi: {procurement.get('route_cut_count', 0)}",
        f"- Fiziksel makara: {procurement.get('drum_count', 0)}",
        f"- Toplam / en büyük aşım: {procurement.get('overload_total_m', 0):.3f} / {procurement.get('maximum_single_drum_overload_m', 0):.3f} m",
        f"- JSON/CSV/XLSX tutarlılığı: {procurement.get('export_consistency')}",
        "",
        "## Tarihsel düzeltme",
        "",
        "`PACKAGED_TEST_RESULTS_v0.16.9.4.14.txt` ve `PUBLISH_CLEANUP_AUDIT_v0.16.9.4.14.md` içindeki PDF ve azami makara boyu PASS beyanları eksiksiz makine-okunur kanıta dayanmıyordu. Bu paket bu beyanları kanıt kabul etmez ve otomatik sayısal denetimle geçersiz kılar.",
        "",
        "_Bu belge `tools/run_release_acceptance.py` tarafından otomatik üretilmiştir._",
    ]
    if audit["blocking_finding_count"]:
        lines.extend(["", "## Engelleyici bulgular", ""])
        for item in audit["findings"]:
            if not item["allowed"]:
                lines.append(f"- `{item['rule_id']}` · `{item['path']}` · {item['surface']} · {item['location']} · `{item['redacted_preview']}`")
    return "\n".join(lines) + "\n"


def _write_outputs(payload: dict[str, Any], json_path: Path, text_path: Path, markdown_path: Path) -> None:
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    text_path.write_text(_render_text(payload), encoding="utf-8")
    markdown_path.write_text(_render_markdown(payload), encoding="utf-8")


def _self_check_outputs(paths: list[Path]) -> str:
    with tempfile.TemporaryDirectory(prefix="ditus-attestation-") as temporary:
        temp_root = Path(temporary)
        for path in paths:
            shutil.copy2(path, temp_root / path.name)
        result = scan_package(temp_root)
        return result.status


def run_release_acceptance(root: Path, version: str, pytest_result: dict[str, Any] | None = None) -> dict[str, Any]:
    root = root.resolve()
    json_name = f"PUBLISH_INTEGRITY_AUDIT_v{version}.json"
    text_name = f"PACKAGED_TEST_RESULTS_v{version}.txt"
    markdown_name = f"PUBLISH_CLEANUP_AUDIT_v{version}.md"
    output_paths = [root / json_name, root / text_name, root / markdown_name]
    output_relatives = {path.relative_to(root).as_posix() for path in output_paths}

    for path in output_paths:
        path.unlink(missing_ok=True)

    pytest_result = pytest_result or run_pytest(root)
    engine_lock = verify_engine_lock(root, root / f"ENGINE_BASELINE_v{version}.sha256")
    audit_result = scan_package(root, excluded_paths=output_relatives | {"MANIFEST.txt"})
    manifest = build_and_verify_manifest(root, root / "MANIFEST.txt", output_relatives)

    procurement_drum_plan = audit_procurement_drum_plan(root)
    payload = _compose_payload(version, pytest_result, audit_result, engine_lock, manifest, procurement_drum_plan, "PENDING")
    _write_outputs(payload, *output_paths)
    self_check = _self_check_outputs(output_paths)
    payload = _compose_payload(version, pytest_result, audit_result, engine_lock, manifest, procurement_drum_plan, self_check)
    _write_outputs(payload, *output_paths)
    final_self_check = _self_check_outputs(output_paths)
    if final_self_check != self_check:
        payload = _compose_payload(version, pytest_result, audit_result, engine_lock, manifest, procurement_drum_plan, final_self_check)
        _write_outputs(payload, *output_paths)
    return payload


def _main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run DiTuS release acceptance and generate attestations")
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--version", required=True)
    parser.add_argument("--shard-count", type=int, default=8)
    parser.add_argument("--run-shard-index", type=int)
    parser.add_argument("--finalize-from-shards", action="store_true")
    args = parser.parse_args(argv)
    root = args.root.resolve()

    if args.run_shard_index is not None:
        result = run_pytest_shard(root, args.run_shard_index, args.shard_count)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result["status"] == "PASS" else 1

    if args.finalize_from_shards:
        pytest_result = aggregate_pytest_shards(root, args.shard_count)
        payload = run_release_acceptance(root, args.version, pytest_result=pytest_result)
    else:
        payload = run_release_acceptance(root, args.version)

    print(json.dumps({
        "overall_status": payload["overall_status"],
        "tests": payload["pytest"]["tests"],
        "passed": payload["pytest"]["passed"],
        "publish_integrity": payload["publish_integrity"]["status"],
        "engine_lock": payload["engine_lock"]["status"],
        "manifest": payload["manifest"]["status"],
    }, ensure_ascii=False, indent=2))
    return 0 if payload["overall_status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(_main())
