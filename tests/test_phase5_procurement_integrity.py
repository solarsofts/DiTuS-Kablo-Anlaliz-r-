from __future__ import annotations

from copy import deepcopy
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from ucd.calculations.bonding_accessories import resolve_bonding_accessory_plan
from ucd.calculations.procurement import (
    DrumCut,
    _pack_drums,
    build_procurement_package,
    write_procurement_package,
)
from ucd.models.project import ProjectData

ROOT = Path(__file__).resolve().parents[1]


def _project(name: str = "synthetic_20km_procurement.ucd.json") -> ProjectData:
    return ProjectData.from_dict(json.loads((ROOT / "examples" / name).read_text(encoding="utf-8")))


def test_phase5_synthetic_drum_plan_distributes_allowance_without_overload() -> None:
    package = build_procurement_package(_project())
    assert package.summary.drum_plan_status == "VALID"
    assert package.summary.drum_count == 126
    assert package.summary.route_cut_total_m == 121752.0
    assert package.summary.order_allowance_total_m == 2423.0
    assert package.summary.overload_total_m == 0.0
    assert package.summary.unallocated_total_m == 0.0
    assert package.summary.unassigned_cut_count == 0
    assert not any(cut.cut_type == "ORDER_RESERVE" for drum in package.drums for cut in drum.cuts)
    assert all(drum.overload_m == 0.0 for drum in package.drums)
    assert all(drum.capacity_balance_m >= 0.0 for drum in package.drums)


def test_phase5_long_route_cut_is_unassigned_and_plan_invalid() -> None:
    cut = DrumCut("CUT-LONG", "R1", 1, "A", 1, 1200.0, 0.0, 0.0, 1250.0)
    drums, unassigned, summary, warnings = _pack_drums(
        [cut], 1000.0, 0.0, "CBL-001", order_quantity_m=1250.0, spare_stock_total_m=0.0
    )
    assert not drums
    assert len(unassigned) == 1
    assert unassigned[0].deficit_m == 250.0
    assert summary.drum_plan_status == "INVALID"
    assert summary.unassigned_cut_count == 1
    assert warnings


def test_phase5_operational_spare_uses_dedicated_bounded_drums() -> None:
    cut = DrumCut("CUT-1", "R1", 1, "A", 1, 500.0, 0.0, 0.0, 500.0)
    drums, unassigned, summary, _ = _pack_drums(
        [cut], 1000.0, 0.0, "CBL-001", order_quantity_m=2923.0, spare_stock_total_m=2423.0
    )
    assert not unassigned
    spare = [drum for drum in drums if drum.spare_stock_length_m > 0]
    assert [drum.spare_stock_length_m for drum in spare] == [1000.0, 1000.0, 423.0]
    assert all(drum.loaded_length_m <= 1000.0 for drum in drums)
    assert summary.spare_stock_total_m == 2423.0


def test_phase5_accessory_plan_derives_major_and_minor_boundaries() -> None:
    project = _project()
    plan = resolve_bonding_accessory_plan(project.bonding)
    assert plan.status == "VALID"
    assert plan.cross_boundary_count == 14
    assert plan.major_ground_boundary_count == 6
    assert plan.cross_link_box_units_per_circuit == 14
    assert plan.grounding_link_box_units_per_circuit == 6
    assert plan.svl_set_units_per_circuit == 14
    assert plan.svl_pole_units_per_circuit == 42


def test_phase5_contains_svl_cannot_override_graph_mismatch() -> None:
    project = _project()
    broken = deepcopy(project)
    major_box = next(box for box in broken.bonding.link_boxes if box.joint_node_id == "J003")
    major_box.contains_svl = True
    plan = resolve_bonding_accessory_plan(broken.bonding)
    assert plan.status == "INVALID"
    assert any("SVL_NOT_REQUIRED_AT_SOLID_GROUND_BOUNDARY" in item for item in plan.errors)


def test_phase5_exports_carry_numeric_overload_and_accessory_fields(tmp_path: Path) -> None:
    package = build_procurement_package(_project())
    paths = write_procurement_package(package, tmp_path, "phase5", ("json", "csv", "xlsx"))
    payload = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert payload["summary"]["overload_total_m"] == 0.0
    assert payload["summary"]["svl_pole_units"] == 84
    with paths["csv_drum_plan"].open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter=";"))
    assert rows and "Aşım m" in rows[0]
    assert sum(float(row["Aşım m"]) for row in rows) == 0.0
    wb = load_workbook(paths["xlsx"], read_only=True)
    headers = [cell.value for cell in next(wb["Makara Planı"].iter_rows(min_row=1, max_row=1))]
    wb.close()
    assert {"Aşım [m]", "Kapasite Bakiyesi [m]", "Sipariş/Fire Payı [m]"}.issubset(headers)
